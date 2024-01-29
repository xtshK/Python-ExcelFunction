import openpyxl
import os 
import platform
import tkinter as tk
import pandas as pd
from tkinter import filedialog, messagebox

## Def the variable
platform_name = platform.system()
username = os.environ.get('USERNAME') or os.environ.get('USER')

column_to_check = ['D', 'G', 'J']
column_name=['Year','Month','Country','Model-Type','Model-Size','Model','Currency','Column1','Model Price','Column3','Forecase Qty']
start_row = 3


if platform_name == "Darwin":
    Removed_Save_path=f'/Users/Desktop/AfterRemove.xlsx'
    Convert_Save_path=f'/Users/Desktop/AfterConvert.xlsx'
else:
    Removed_Save_path = f'C:\\4M formatted result\\AfterRemove.xlsx'
    Convert_save_path = f'C:\\4M formatted result\\AfterConvert.xlsx'
    
def center_window(window, width, height):
    screen_width = window.winfo_screenwidth()
    screen_height = window.winfo_screenheight()
    
    x=(screen_width - width)//2
    y=(screen_height - height)//2
    
    window.geometry(f"{width}x{height}+{x}+{y}")    
    
def merage_fun():    
    folder_path=filedialog.askdirectory(title="Selecet the AfertFill_Allocate folder.")
    
    if folder_path:    
        file_paths=[os.path.join(folder_path, file) for file in os.listdir(folder_path)]
        
        
        for index, file_path in enumerate(file_paths, start=1):
                        
            file_name, file_extension = os.path.splitext(os.path.basename(file_path))
            
            new_file_name=f"{file_name}_Meraged{file_extension}"
            new_file_path=os.path.join(folder_path, new_file_name)        
            
                    
            workbook=pd.read_excel(file_path,sheet_name=None,header=None,names=column_name)
            merage_data=pd.concat(workbook.values(), ignore_index=True)
            merage_data.to_excel(new_file_path, index=False)
            
            messagebox.showinfo("Hint",f"{new_file_name}--Finished!!!")
            
    else:
        print('Folder selection canceled.')

def remove_spaces_fun():
    file_path = filedialog.askopenfilename(title="Select the FileInOne_Result file.")
    
    if file_path:
        Workbook = openpyxl.load_workbook(file_path, data_only=True)

        for sheet in Workbook.sheetnames:
            current_sheet = Workbook[sheet]
            
            row_to_delete = [] # 用來存放要刪除的行數

            for row_number, row in enumerate(current_sheet.iter_rows(min_row=3, max_col=1, max_row=current_sheet.max_row, values_only=True), start=3):
                for idx, cell_value in enumerate(row):
                    if cell_value is not None and isinstance(cell_value, str):
                        current_sheet.cell(row=row_number, column=idx+1, value=cell_value.replace(' ', ''))

                if not any(row):
                    row_to_delete.append(row_number)
            for row_number in reversed(row_to_delete):
                current_sheet.delete_rows(row_number)
                
        Workbook.save(Removed_Save_path)
        Workbook.close()
        messagebox.showinfo("Hint","Finished!!!")
        
    else:
        print("File selection canceled.")

def Convert_fun():
    file_path = filedialog.askopenfilename(title="Select the file to convert to values.")

    if file_path:
        Workbook = openpyxl.load_workbook(file_path, data_only=True)
        for sheet in Workbook.sheetnames:
            current_sheet = Workbook[sheet]
            for row in current_sheet.iter_rows(values_only=True):
                for cell_value in row:
                    if isinstance(cell_value, str) and cell_value.startswith('='):
                        try:
                            cell_address = current_sheet.cell(row=row[0], column=row.index(cell_value) + 1).coordinate
                            result = sheet[cell_address].value
                            current_sheet[cell_address] = result
                        except Exception as e:
                            print(f"Unable to evaluate cell {cell_address}: {e}")
        Workbook.save(Convert_save_path)
        Workbook.close()  # Close the workbook
        messagebox.showinfo("Hint","Finished!!!")
        
    else:
        print("File selection canceled.")

def cancel_btn():
    root.destroy()


root = tk.Tk()
root.title("File Dialog")
root.geometry('270x380')

font_style = ('Helvetica', 12, 'bold')

label=tk.Label(root,text='Select a function', font=font_style)
label.pack(pady=15)

label_remove=tk.Label(root,text='1. Select the file \n which you want to remove spaces.')
label_remove.pack(pady=2)

removeBtn=tk.Button(root, text="Remove Spaces",command=remove_spaces_fun,width=13)
removeBtn.pack(pady=15)

label_Convert=tk.Label(root,text='2.Convert cells from "Formula" to "Value".')
label_Convert.pack(pady=2)

ConvertBtn=tk.Button(root, text="Convert File",command=Convert_fun,width=13)
ConvertBtn.pack(pady=15)

label_Merage=tk.Label(root,text='3.Combine the Allocate Files \n worksheets into one worksheet.')
label_Merage.pack(pady=2)

meragrBtn=tk.Button(root, text='Merage File',command=merage_fun,width=13)
meragrBtn.pack(pady=15)

cancelBtn=tk.Button(root, text='Exit', command=cancel_btn, width=5)
cancelBtn.pack(pady=10)

root.mainloop()