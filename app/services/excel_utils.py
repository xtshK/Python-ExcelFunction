import openpyxl
from pathlib import Path
import pandas as pd
import  re

_EXTRA_SPACES = "\u00A0\u2007\u202F\u3000"  # 常見非斷行/全形空白

def _is_blank_a_cell(value) -> bool:
    """判斷 A 欄是否空白（包含全形／NBSP 空白）"""
    if value is None:
        return True
    if isinstance(value, str):
        # 把特殊空白轉成一般空白後去掉
        for ch in _EXTRA_SPACES:
            value = value.replace(ch, " ")
        return value.strip() == ""
    return False

def unique_path(path: Path) -> Path:
    """若檔名存在，自動加 _1, _2… 避免覆蓋"""
    if not path.exists():
        return path
    i = 1
    while True:
        candidate = path.with_name(f"{path.stem}_{i}{path.suffix}")
        if not candidate.exists():
            return candidate
        i += 1

def merge_many_files(folder: str, column_names=None):
    """
    對資料夾中的每個 Excel 檔，將「所有工作表」縱向合併成一張表，
    並輸出回同資料夾，檔名：<原名>_Merged.<ext>
    """
    results = []
    folder_p = Path(folder)
    for p in folder_p.iterdir():
        if p.is_file() and p.suffix.lower() in (".xlsx", ".xlsm", ".xls"):
            df_map = pd.read_excel(p, sheet_name=None, header=None,
                                   names=column_names if column_names else None)
            merged = pd.concat(df_map.values(), ignore_index=True)
            out = p.with_name(f"{p.stem}_Merged{p.suffix}")
            merged.to_excel(out, index=False)
            results.append(out)
    return results


def remove_spaces_first_col_and_drop_blank_rows(src_path: str, out_path: str, start_row: int = 3):
    """
    從第 start_row 列開始：
      1) 移除 A 欄內的所有空白字元（含全形／NBSP）
      2) 若 A 欄為空白，則整列刪除
    存成 out_path
    """
    wb = openpyxl.load_workbook(src_path, data_only=True)

    for name in wb.sheetnames:
        ws = wb[name]
        to_delete = []

        for r in range(start_row, ws.max_row + 1):
            cell = ws.cell(row=r, column=1)
            # 1️⃣ 先清理空白
            if isinstance(cell.value, str):
                for ch in _EXTRA_SPACES:
                    cell.value = cell.value.replace(ch, " ")
                cell.value = cell.value.replace(" ", "")

            # 2️⃣ 若 A 欄是空白 → 記錄刪除
            if _is_blank_a_cell(cell.value):
                to_delete.append(r)

        # 反向刪除，避免位移錯亂
        for r in reversed(to_delete):
            ws.delete_rows(r)

    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(out))
    wb.close()
    return str(out)

def convert_formulas_to_values(src_path: str, out_path: str):
    """
    將公式覆寫為其「快取值」：
    - 用 data_only=True 讀可見值
    - 再寫回原工作簿結構
    注意：若檔案從未被 Excel 計算/儲存過，快取值可能為 None
    """
    wb_values = openpyxl.load_workbook(src_path, data_only=True)
    wb_write  = openpyxl.load_workbook(src_path, data_only=False)

    for name in wb_write.sheetnames:
        ws_w = wb_write[name]
        ws_v = wb_values[name]
        max_row = max(ws_w.max_row, ws_v.max_row)
        max_col = max(ws_w.max_column, ws_v.max_column)
        for r in range(1, max_row + 1):
            for c in range(1, max_col + 1):
                ws_w.cell(row=r, column=c).value = ws_v.cell(row=r, column=c).value

    out = Path(out_path)
    out.parent.mkdir(parents=True, exist_ok=True)
    wb_write.save(str(out))
    wb_values.close()
    wb_write.close()
    return str(out)

def clean_excel_file_to_values(src_path: str, out_dir: str | None = None) -> str:
    """
    讀取所有工作表為當前值（不含公式計算），輸出 *_cleaned.xlsx。
    注意：若原檔從未被 Excel 開啟/儲存過，pandas/openpyxl 讀到的可能是空值（沒快取）。
    """
    src = Path(src_path)
    engine = "openpyxl" if src.suffix.lower() == ".xlsx" else "xlrd"  # .xls 需 xlrd==1.2.0
    excel_data = pd.read_excel(src, sheet_name=None, engine=engine)

    out_dir_p = Path(out_dir) if out_dir else src.parent
    out_dir_p.mkdir(parents=True, exist_ok=True)

    # 產生不覆蓋的檔名
    base = src.stem
    out = out_dir_p / f"{base}_cleaned.xlsx"
    i = 1
    while out.exists():
        out = out_dir_p / f"{base}_cleaned_{i}.xlsx"
        i += 1

    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        for sheet_name, df in excel_data.items():
            safe = str(sheet_name)[:31]
            for ch in "[]*?:/\\":
                safe = safe.replace(ch, "_")
            df.to_excel(writer, sheet_name=safe, index=False)

    return str(out)